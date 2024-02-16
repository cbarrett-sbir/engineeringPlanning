import os
import datetime as dt
import argparse

import openpyxl as opxl
import pandas as pd
import click

from tests import testContractValidity, isValidName, testSheetExistence, isCorrectDate, weekContractsMatch
from fileIO import getDefaultPaths, getContractList, getTeamList

class Person:
    def __init__(self, name="", forecast_date="", schedule_type="9/80", alternate_hours=None, contracts=None):
        """
        Initialize a Person object.

        Parameters:
        - name (str): The name of the person (default is an empty string).
        - forecast_date (str): The time forecast date (e.g., "01/31/2024", default is an empty string).
        - schedule_type (str): The schedule type ("9/80" or "40 hours", default is "9/80").
        - alternate_hours (tuple): A tuple to record alternate 8 hours (e.g., (True, 1) for week 1, default is None).
        - contracts (pd.Series): A pandas Series of contracts (default is None).
        """
        self.name = name
        self.forecast_date = forecast_date
        self.schedule_type = schedule_type
        self.alternate_hours = alternate_hours
        self.contracts = contracts

    def __str__(self):
        return f"Person(name={self.name}, forecast_date={self.forecast_date}, schedule_type={self.schedule_type}, alternate_hours={self.alternate_hours}, contracts={self.contracts})"


# silence obnoxious false positive warning
# default='warn'
pd.options.mode.chained_assignment = None

if __name__ == "__main__":
    # Parse command-line arguments
    parser = argparse.ArgumentParser(description='Check a directory of time forecast excel sheets for correct names, dates, contracts, etc...')
    parser.add_argument('--all', action='store_true', help='Print all test results to report file (else only failing tests show)')
    args = parser.parse_args()

    # - check if name is valid
    # - check if date is correct
    # - add name to a list of names so we can check
    # who is missing at the end

    # - check if contracts are in list
    # - determine which schedule they are on
    # and check if they have populated alternate hours

    if args.all:
        print("CALLED WITH ALL")

    DEFAULTS = r"Q:\EngineeringPlanning\ReportTools\defaults\config.json"
    print(f"Fetching defaults from: {DEFAULTS}")
    tf, out, CN_LIST_PATH, TEAM_LIST_PATH = getDefaultPaths(DEFAULTS)

    SHEETS = click.prompt(
        "Path to the directory containing time forecast excel sheets",
        type=str,
        default=tf
        )
    OUTPUT = click.prompt(
        "Path to place the report",
        type=str,
        default=out
        )

    while True:
        user_input = input("Enter the correct week beginning date (MM/DD/YYYY): ")

        try:
            # Parse the input date
            week_begin = dt.datetime.strptime(user_input, "%m/%d/%Y")
        except ValueError:
            print("Invalid date format. Please use the format month/day/year.")
            continue
        break

    print(f"You entered: {week_begin}")

    CONTRACT_LIST = getContractList(CN_LIST_PATH)
    team_list = getTeamList(TEAM_LIST_PATH)

    current_time = dt.datetime.now().strftime("%Y-%m-%d--%H-%M-%S")

    # Write the report content to the file
    with open(os.path.join(OUTPUT, f"validation_report_{current_time}.txt"), 'w', encoding='UTF-8') as file:
        header = f"Time Forecast Data Validation Report\nGENERATED: {current_time}\nFor week beginning: {week_begin}"
        file.write(header)

        present_names = []
        for filename in os.listdir(SHEETS):
            if filename.endswith(".xlsm"):
                if (filename.startswith('~')):
                    print(f"Temporary file detected: {filename}")
                    continue

                error = f"\n\nEvaluating {filename}..."
                
                print(f"\nEvaluating {filename}...")
                file_path = os.path.join(SHEETS, filename)

                wb = opxl.load_workbook(file_path, read_only=True, data_only=True)
                testSheetExistence(wb)

                ws = wb["Plan"]

                team_member = Person()
                team_member.name = ws.cell(6, 5).value
                team_member.forecast_date = ws.cell(7, 5).value

                sch = ws.cell(2, 1).value # 1 = 9/80, 2 = 40 hr
                if (sch == 1):
                    team_member.schedule_type = "9/80"
                else:
                    team_member.schedule_type = "40"

                # need to say "or 0" in case
                team_member.alternate_hours = int(ws.cell(39, 11).value or 0) + int(ws.cell(39, 24).value or 0)
                team_member.contracts = pd.read_excel(
                    io=file_path,
                    sheet_name="Plan",
                    header=None,
                    usecols="C",
                    skiprows=range(0, 17),
                    nrows=14,
                    keep_default_na=True,
                    names=["contract"]
                ).dropna()["contract"]

                # Name testing
                if isValidName(team_member.name, team_list["name"]):
                    print("PASSED: Name Validity")
                    present_names.append(team_member.name)
                else:
                    s = f"FAILED: Name Validity, {team_member.name} is not in team member list!"
                    print(s)
                    error += '\n' + f"FAILED: Name Validity, {team_member.name} is not in team member list!"

                # Date testing
                if isCorrectDate(week_begin, team_member.forecast_date):
                    print("PASSED: Date Correctness")
                else:
                    s = f"FAILED: Date Correctness, {team_member.forecast_date.strftime('%Y-%m-%d')} != {week_begin.strftime('%Y-%m-%d')} (actual)!"
                    print(s)
                    error += '\n' + s

                # Contract testing
                if not team_member.contracts.empty:
                    mismatched_cn = weekContractsMatch(file_path, team_member.contracts)
                    if not mismatched_cn:
                        print("PASSED: Week 1 == week 2 contracts")
                    else:
                        s = f"FAILED: Week 1 != week 2 contracts, missing contracts: {mismatched_cn}"
                        print(s)
                        error += '\n' + s

                    xs = testContractValidity(team_member.contracts, CONTRACT_LIST["contract"])
                    if (xs == []):
                        print("PASSED: Contract Validity")
                    else:
                        s = "WARNING: Contracts included but not found in ContractList.xlsx:"
                        for x in xs:
                            s += '\n ' + str(x)
                        print(s)
                        error += '\n' + s
                else:
                    s = f"WARNING: {team_member.name} did not enter any contracts"
                    print(s)
                    error += '\n' + s

                # Schedule testing
                if team_member.schedule_type == "9/80":
                    if team_member.alternate_hours > 0:
                        print("PASSED: 9/80 Alt Hours")
                    else:
                        s = f"FAILED: 9/80 Alt Hours, {team_member.name} works 9/80 but entered 0 alt hours"
                        print(s)
                        error += '\n' + s

                if error != f"\n\nEvaluating {filename}...":
                    file.write(error)
                    

        print("\nReports missing:")
        error = "\n\nReports missing:"
        for n in team_list["name"][~team_list["name"].isin(present_names)].tolist():
            print(n)
            error += '\n' + n
        
        file.write(error)