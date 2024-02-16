import datetime

import openpyxl
import pandas as pd
import click

from fileIO import retrieveTimeForecasts, printHeader, dataframeToExcel, getDefaultPaths, getTeamList
from manipulate import filterNaNs

# silence obnoxious false positive warning
# default='warn'
pd.options.mode.chained_assignment = None

if __name__ == "__main__":
    # The high-level algorithm runs as follows:
    #   1. Gather all time forecast data into dataframe
    #      "forecasts" (pd.Dataframe)
    #
    #   2. Retrieve team member list (which contains
    #      team member -> group mapping) to organize
    #      team members by their discipline (e.g. ME, EE, ...)
    #
    #   3. Match team member data to their group labels
    #
    #   4. Iterate through active disciplines
    #       4.1 Get names present in an active discipline
    #           4.1.1 Get contracts worked on by each name
    #                 and print them to excel
    #
    #   5. Format excel sheet

    DEFAULTS = r"Q:\EngineeringPlanning\ReportTools\defaults\config.json"
    print(f"Fetching defaults from: {DEFAULTS}")
    tf, out, cn, tl = getDefaultPaths(DEFAULTS)

    SHEETS = click.prompt(
        "Path to the directory containing time forecast excel sheets",
        type=str,
        default=tf
        )
    OUTPUT = click.prompt(
        "Path to place the report",
        type=str,
        default=out
        )
    CN_LIST_PATH = cn
    TEAM_LIST_PATH = tl

    forecasts, DATE = retrieveTimeForecasts(SHEETS)
    forecasts = filterNaNs(forecasts)

    print("Reading TeamMembersList.xlsx...")
    try:
        team_list = pd.read_excel(TEAM_LIST_PATH, "Sheet1", header=0)
    except Exception as e:
        print(e)

    col_names = ['name', 'group', 'group_list', 'manager']
    team_list = team_list.set_axis(col_names, axis='columns')
    '''
    ex: team_list
    ----------------------------------------
    name        group   group_list  manager
    Al Gibson   SRS     SE          Jim M
    Anita Chow  SWE     IRSP        Tom D
    ...
    ----------------------------------------
    '''

    print("Reading ContractList.xlsx...")
    try:
        contract_list = pd.read_excel(CN_LIST_PATH, "Sheet1", header=None)
    except Exception as e:
        print(e)
    
    # create basis for contract info dataframe
    contract_list = contract_list[[0, 1]][1:]
    contract_list = contract_list.set_axis(['contract','desc'], axis='columns')

    print("Matching contracts and descriptions...")
    forecasts = pd.merge(
        forecasts, 
        contract_list[['contract', 'desc']], 
        on="contract", how='left'
    )

    print("Fetching a list of disciplines...")
    # create a list of disciplines to iterate
    # through when printing the report
    disciplines = team_list.iloc[:, 2].dropna().tolist()
    '''
    ex: disciplines
    ----------
    0      SE
    1    IRSP
    2     SWE
    3      ME
    4      EE
    5     SRS
    ----------
    '''

    print("Matching disciplines to people...")
    # add discipline column to forecasts data
    forecasts = pd.merge(
        left=forecasts,
        right=team_list[["name", "group"]],
        on="name",
        how='left'
    )

    # replace blank (NaN) contracts with "none" for grouping. 
    # (np.NaN cannot be passed as key to get_group)
    values = {"contract": "none"}
    forecasts.fillna(value=values, inplace=True)

    # reorder cols to move desc to col 2
    forecasts = forecasts[['name', 'week', 'contract', 'desc', 'monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'roll_up_hours', 'roll_up_percent', 'milestone1', 'milestone2', 'milestone3', 'group']]

    # cut rows with unallocated time = 0
    forecasts = forecasts[~((forecasts['contract'] == 'Unallocated Time') & (forecasts['roll_up_hours'] == 0))]


    # see which disciplines are present for this 2-week period
    active_disciplines = forecasts["group"].unique()

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Report", 0)  # insert at first position

    H = ["Name",
         "Week",
         "Contract",
         "Description",
         "M", "T", "W", "R", "F",
         "Hours",
         "%",
         "Milestone 1", "Milestone 2","Milestone 3"]

    # group all forecast rows by discipline for printing
    discipline_groups = forecasts.groupby(['group'])

    # tracks current row being printed to excel sheet
    curr_row = 1
    printHeader(
        ws,
        curr_row,
        [f"REPORT FOR WEEK BEGINNING: {str(DATE)}, "
          f"GENERATED: {datetime.datetime.now()}"]
    )
    curr_row += 1

    printHeader(ws, curr_row, H)
    curr_row += 1

    num_active_disciplines = 0  # flag to avoid printing header twice
    # only print header again if there is > 1 dicipline active
    for i, discipline in enumerate(disciplines):
        # if discipline in data, fetch their associated contracts
        if discipline in active_disciplines:
            print(f"Writing report for {discipline}...")
            num_active_disciplines += 1
        else:
            continue

        if num_active_disciplines > 1:
            printHeader(ws, curr_row, H)
            curr_row += 1

        # get names present in this disciplines data
        names = forecasts.iloc[discipline_groups.indices.get((discipline))
                               ]['name'].unique()

        for name in names:
            # track first row index for use vertically appending later
            FIRST_ROW = curr_row

            # write persons name on left and their group above week #
            person_info = pd.DataFrame({'name':[name], 'group':[discipline]})
            dataframeToExcel(person_info, ws, False, False, False, curr_row, 1)
            curr_row += 1

            s = forecasts.groupby(["group", "name", "week"])

            week1 = forecasts.iloc[s.indices.get((discipline, name, 1))]
            week2 = forecasts.iloc[s.indices.get((discipline, name, 2))]

            '''
            ex: week1
            -----------------------------------------
            name        week    contract    monday
            Jack Grigor     1   82500/DOC      1  ...     
            Jack Grigor     1   82500/TST      8  ...  
            Jack Grigor     1  Sustaining      1  ...
            -----------------------------------------
            '''

            # drop discipline col to avoid it being printed
            # at the end of the row entry
            week1.drop(columns=['group'], inplace=True)
            week2.drop(columns=['group'], inplace=True)

            dataframeToExcel(week1, ws, False, False, True, curr_row, 1)
            curr_row += len(week1.index)
            week1_length = len(week1.index)

            dataframeToExcel(week2, ws, False, False, False, curr_row, 1)
            curr_row += len(week2.index) + 1
            week2_length = len(week2.index)

            # merge contract cells
            ws.merge_cells(
                start_row=FIRST_ROW,
                start_column=1,
                end_row=FIRST_ROW + week1_length + week2_length,
                end_column=1
            )

            # merge week number cells
            ws.merge_cells(
                start_row=FIRST_ROW + 1,
                start_column=2,
                end_row=FIRST_ROW + week1_length,
                end_column=2
            )
            start_week2 = FIRST_ROW + week1_length + 1

            ws.merge_cells(
                start_row=start_week2,
                start_column=2,
                end_row=start_week2 + week2_length - 1,
                end_column=2
            )

    print("Formatting...")
    # [1:] is to avoid centering report date
    # in first cell
    for cell in ws['A'][1:]:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center'
        )

    for cell in ws['B']:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center'
        )

    # avoid spill over on description column
    for cell in ws['D']:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='fill',
        )

    WEEKDAY_WIDTH = 5
    ws.column_dimensions['A'].width = 19    # name
    ws.column_dimensions['B'].width = 6     # week
    ws.column_dimensions['C'].width = 11    # contract
    ws.column_dimensions['D'].width = 19    # desc
    ws.column_dimensions['E'].width = WEEKDAY_WIDTH  # m
    ws.column_dimensions['F'].width = WEEKDAY_WIDTH  # t
    ws.column_dimensions['G'].width = WEEKDAY_WIDTH  # w
    ws.column_dimensions['H'].width = WEEKDAY_WIDTH  # r
    ws.column_dimensions['I'].width = WEEKDAY_WIDTH  # f
    ws.column_dimensions['J'].width = 6  # hours
    ws.column_dimensions['K'].width = WEEKDAY_WIDTH + .5  # %
    ws.column_dimensions['L'].width = 13  # milestone 1
    ws.column_dimensions['M'].width = 13  # milestone 2
    ws.column_dimensions['N'].width = 13  # milestone 3

    # style weekday and hours columns
    for row in ws.iter_rows(min_row=1, min_col=4, max_col=9):
        for cell in row:
            cell.number_format = "0.0"
    
    # style % column
    for row in ws.iter_rows(min_row=1, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = "0%"

    printHeader(ws, curr_row, ["Team Members Reported:"])
    curr_row += 1
    unique_names = sorted(forecasts["name"].unique())
    for name in unique_names:
        printHeader(ws, curr_row, [name])
        curr_row += 1

    curr_row += 1
    printHeader(ws, curr_row, ["Members Missing:"])
    curr_row += 1
    team_list = getTeamList(TEAM_LIST_PATH)
    for n in team_list["name"][~team_list["name"].str.lower().isin(list(name.lower() for name in unique_names))].tolist():
        printHeader(ws, curr_row, [n])
        curr_row += 1

    print("Saving...")

    wb.save(OUTPUT + "/Team_Report_for_" + str(DATE) + ".xlsx")

    print("Report compiled successfully!")
    print()
    input("Press Enter to quit...")
