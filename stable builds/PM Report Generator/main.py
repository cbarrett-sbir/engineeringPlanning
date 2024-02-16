import datetime

import openpyxl
import pandas as pd
import click

from fileIO import retrieveTimeForecasts, printHeader, dataframeToExcel, getDefaultPaths, getTeamList
from manipulate import filterNaNs

# silence obnoxious false positive warning
# default='warn'
pd.options.mode.chained_assignment = None

if __name__ == "__main__":
    # The high-level algorithm runs as follows:
    #   1. Gather all time forecast data into dataframe
    #      "forecasts" (pd.Dataframe)
    #
    #   2. Retrieve contract list (which contains
    #      contract -> name) to organize people
    #      by the contracts they will work on
    #
    #   3. Match team member data to their contract
    #
    #   4. Iterate through active contracts
    #       4.1 Get names present in an active contract
    #           4.1.1 Print names working on the contract
    #                 to excel
    #
    #   5. Format excel sheet
    
    DEFAULTS = r"Q:\EngineeringPlanning\ReportTools\defaults\config.json"
    print(f"Fetching defaults from: {DEFAULTS}")
    tf, out, cn, tl = getDefaultPaths(DEFAULTS)

    SHEETS = click.prompt(
        "Path to the directory containing time forecast excel sheets",
        type=str,
        default=tf
        )
    OUTPUT = click.prompt(
        "Path to place the report",
        type=str,
        default=out
        )
    CN_LIST_PATH = cn
    TEAM_LIST_PATH = tl

    forecasts, DATE = retrieveTimeForecasts(SHEETS)
    forecasts = filterNaNs(forecasts)

    print("Reading ContractList.xlsx...")
    try:
        contract_list = pd.read_excel(CN_LIST_PATH, "Sheet1", header=None)
    except Exception as e:
        print(e)

    # create list of program managers from ContractsList sheet
    print("Fetching a list of active program managers...")
    program_mgr_list = contract_list[[4]][2:]
    program_mgr_list.columns = ["program_mgr"]
    program_mgr_list.dropna(inplace=True)

    # create basis for contract info dataframe
    contract_list = contract_list[[0, 2, 1]][1:]
    contract_list = contract_list.set_axis(['contract', 'program_mgr', 'desc'], axis='columns')


    # merge associated program mgr labels to contracts and sort
    print("Matching managers to contracts...")
    contracts_with_pm = pd.merge(
        forecasts,
        contract_list[['contract', 'program_mgr']],
        on="contract", how='left'
    )
    contracts_with_pm = contracts_with_pm.sort_values(["contract", "week", "name"])
    
    # replace NaNs with "none" for grouping. (np.NaN cannot be passed as key to get_group)
    values = {"program_mgr":"none", "contract":"none"}
    contracts_with_pm.fillna(value=values, inplace=True)

    # create an ordered list of PMs based on their order in ContractsList.xlsx
    mgr_order = program_mgr_list['program_mgr'].tolist()
    # used to aggregates NaNs last in print out (np.NaN cannot be passed as key to get_group)
    mgr_order.append("none")

    # see which PMs and contracts are present for this 2 week period
    active_mgrs = contracts_with_pm["program_mgr"].unique()

    # split data by PM
    mgr_groups = contracts_with_pm.groupby(["program_mgr"])




    wb = openpyxl.Workbook()
    ws = wb.create_sheet("Report", 0) # insert at first position

    # tracks current row being printed to excel sheet
    curr_row = 1
    printHeader(
        ws,
        curr_row,
        [f"REPORT FOR WEEK BEGINNING: {str(DATE)}, "
         f"GENERATED: {datetime.datetime.now()}"]
    )
    curr_row += 1

    H = ["Contract",
        "Week",
        "Name",
        "M", "T", "W", "R", "F",
        "Hours",
        "%",
        "Milestone 1", "Milestone 2","Milestone 3"]

    printHeader(ws, curr_row, H)
    curr_row += 1

    num_active_managers = 0 # flag to avoid printing header twice
    # only print header again if there is > 1 mgr active
    for i, mgr in enumerate(mgr_order):
        # if manager in data, fetch their associated contracts
        if mgr in active_mgrs:
            group = mgr_groups.get_group((mgr))
            group.drop(columns=['program_mgr'], inplace=True)
            print(f"Writing report for {mgr}...")
            num_active_managers += 1
        else:
            continue

        if num_active_managers > 1:
            printHeader(ws, curr_row, H)
            curr_row += 1
            
        # get contracts present in this mgrs data
        contracts = group["contract"].unique()
        contract_groups = group.groupby(["contract", "week"])

        for contract in contracts:
            # track first row index for use vertically appending later
            FIRST_ROW = curr_row

            contract_info = contract_list[contract_list["contract"] == contract]
            if (contract not in contract_list.contract.values): # create new info so the weeks will have a CN label
                contract_info = pd.DataFrame({'contract':[contract]})
                if (contract not in ["Sustaining", "ENG_OH", "IRC_OH", "STE_OH", "BP", "PTO", "HOLIDAY"]):
                    name = {contract_groups.get_group((contract, 1)).name.values[0]}
                    print(f"Contract \"{contract}\" was referenced by {name}, but not found in ContractList.xlsx")
            dataframeToExcel(contract_info, ws, False, False, False, curr_row, 1)
            curr_row += 1

            week1 = contract_groups.get_group((contract, 1))
            week2 = contract_groups.get_group((contract, 2))

            dataframeToExcel(week1, ws, False, False, True, curr_row, 1)  
            curr_row += len(week1.index)
            week1_length = len(week1.index)

            dataframeToExcel(week2, ws, False, False, False, curr_row, 1)
            curr_row += len(week2.index) + 1
            week2_length = len(week2.index)

            # track last row index for use merging later
            last_row = curr_row - 2

            # merge contract cells
            ws.merge_cells(
                start_row=FIRST_ROW,
                start_column=1,
                end_row=FIRST_ROW + week1_length + week2_length,
                end_column=1
            )

            # merge week number cells
            ws.merge_cells(
                start_row=FIRST_ROW + 1,
                start_column=2, 
                end_row=FIRST_ROW + week1_length, 
                end_column=2
            )
            start_week2 = FIRST_ROW + week1_length + 1

            ws.merge_cells(
                start_row=start_week2, 
                start_column=2, 
                end_row=start_week2 + week2_length - 1,
                end_column=2
            )

    print("Formatting...")
    # [1:] is to avoid centering report date
    # in first cell
    for cell in ws['A'][1:]:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center'
        )

    for cell in ws['B']:
        cell.alignment = openpyxl.styles.Alignment(
            horizontal='center',
            vertical='center'
        )

    WEEKDAY_WIDTH = 5
    ws.column_dimensions['A'].width = 11  # contract
    ws.column_dimensions['B'].width = 6  # week
    ws.column_dimensions['C'].width = 19  # name
    ws.column_dimensions['D'].width = WEEKDAY_WIDTH  # m
    ws.column_dimensions['E'].width = WEEKDAY_WIDTH  # t
    ws.column_dimensions['F'].width = WEEKDAY_WIDTH  # w
    ws.column_dimensions['G'].width = WEEKDAY_WIDTH  # r
    ws.column_dimensions['H'].width = WEEKDAY_WIDTH  # f
    ws.column_dimensions['I'].width = 6  # hours
    ws.column_dimensions['J'].width = WEEKDAY_WIDTH + .5  # %
    ws.column_dimensions['K'].width = 13  # milestone 1
    ws.column_dimensions['L'].width = 13  # milestone 2
    ws.column_dimensions['M'].width = 13  # milestone 3

    # style weekday and hours columns
    for row in ws.iter_rows(min_row=1, min_col=4, max_col=9):
        for cell in row:
            cell.number_format = "0.0"
    # style % column
    for row in ws.iter_rows(min_row=1, min_col=10, max_col=10):
        for cell in row:
            cell.number_format = "0%"

    printHeader(ws, curr_row, ["Team Members Reported:"])
    curr_row += 1
    unique_names = sorted(forecasts["name"].unique())
    for name in unique_names:
        printHeader(ws, curr_row, [name])
        curr_row += 1

    curr_row += 1
    printHeader(ws, curr_row, ["Members Missing:"])
    curr_row += 1
    team_list = getTeamList(TEAM_LIST_PATH)
    for n in team_list["name"][~team_list["name"].str.lower().isin(list(name.lower() for name in unique_names))].tolist():
        printHeader(ws, curr_row, [n])
        curr_row += 1

    print("Saving...")

    wb.save(OUTPUT + "/PM_Report_for_" + str(DATE) + ".xlsx")

    print("Report compiled successfully!")
