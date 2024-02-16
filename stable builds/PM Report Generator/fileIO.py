'''
The module provides utility functions
for reading excel data into pandas dataframes
and writing those dataframes back to excel
'''
import os
import sys
import warnings
import datetime
from typing import List, Tuple
import json

import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

def excelToDataframe(filepath: str) -> Tuple[pd.DataFrame, datetime.date]:
    '''
    Reads time forecast formatted excel file 
    into a labeled pandas dataframe.

    Params
    ------
        filepath: a valid fielpath to an .xlsx file 
        formatted in the time forecast sheet pattern

    Returns
    -------
        date: week beginning date from sheet
        data: dataframe with following columns
        ---------------
        name             
        week
        contract
        monday
        tuesday
        wednesday
        thursday
        friday
        roll_up_hours
        roll_up_percent
        milestone1
        milestone2
        milestone3
        ---------------
    '''

    # Check if the file is a valid Excel file
    try:
        with warnings.catch_warnings():
            print(f"Loading {filepath}")
            warnings.filterwarnings("ignore", category=UserWarning)
            df = pd.read_excel(filepath, "Plan", header=None)
    except Exception as e:
        print(e)
        return
    
    DATE = df.iloc[6,4].date()

    # Extract name from the DataFrame
    name = df.iloc[5,4]

    # Using iloc for integer-location based indexing
    week1_column_idx = [2, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
    week2_column_idx = [18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28]
    week1 = df.iloc[17:38, week1_column_idx].dropna(how='all', axis=0)
    week2 = df.iloc[17:38, week2_column_idx].dropna(how='all', axis=0)

    # Define column names
    column_names = [
        "contract", 
        "monday", 
        "tuesday", 
        "wednesday", 
        "thursday", 
        "friday", 
        "roll_up_hours", 
        "roll_up_percent", 
        "milestone1", 
        "milestone2", 
        "milestone3"
    ]

    week1.columns = column_names
    week2.columns = column_names

    # Add name and week columns
    week1.insert(1, "name", name)
    week1.insert(1, "week", 1)
    week2.insert(1, "name", name)
    week2.insert(1, "week", 2)

    # Concatenate DataFrames
    data = pd.concat([week1, week2], ignore_index=True)

    return data, DATE

def retrieveTimeForecasts(path: str) -> Tuple[pd.DataFrame, datetime.date]:
    '''
    Returns a pandas dataframe containing all 
    the time report information by contract.

    Params
    ------
        path: path to a directory containing excel 
        sheets with bi-weekly time forecasts

    Returns
    -------
        DATE: week beginning date from sheet
        data: multiple entries of format
        ------------------------------------------
        name             Shaun Reed
        week             1
        contract         15033/905
        monday           4
        tuesday          1
        wednesday        1
        thursday         1
        friday           1
        roll_up_hours    8
        roll_up_percent  0.181818
        milestone1       finish ECO 061590
        milestone2       support whatever pops up
        milestone3       NaN
        ------------------------------------------
    '''

    # Check if the entered path is a valid directory
    if not os.path.isdir(path):
        print("Invalid time forcast directory path.")
        sys.exit()

    # Initialize an empty DataFrame to store the combined data
    data = pd.DataFrame()

    # Iterate through all sheets in the directory, 
    # appending data to dataframe
    for filename in os.listdir(path):
        if filename.endswith(".xlsm"):
            if (filename.startswith('~')):
                print(f"Ignoring temporary file: {filename}")
                continue
            # Construct the full file path
            file_path = os.path.join(path, filename)

            # Convert each Excel file to dataframe 
            forecast, date = excelToDataframe(file_path)

            # Concatenate the result to the data DataFrame
            data = pd.concat([data, forecast], ignore_index=True)

    return data, date

def printHeader(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        header: List
) -> None:
    '''
    Prints a header across the row of a worksheet in bold.

    Params
    ------
        ws: an open worksheet where the header is printed
        row: row number (1-indexed) of worksheet on which to print
        (placement will always begin from the first column)
        header: list of strings to print
    '''
    for idx, item in enumerate(header):
        ws.cell(column=idx + 1, row=row, value=item)
        ws.cell(column=idx + 1, row=row).font = openpyxl.styles.Font(bold=True)

blue_fill = openpyxl.styles.PatternFill('solid', fgColor="daeef3")
style = openpyxl.styles.Side(border_style="hair")
border_style = openpyxl.styles.borders.Border(
    left=style, 
    right=style, 
    top=style, 
    bottom=style,
)

def dataframeToExcel(df: pd.DataFrame,
                     ws: openpyxl.worksheet,
                     header: bool =True,
                     index: bool =True,
                     color: bool =False,
                     startrow: int =0,
                     startcol: int =0
    ) -> None:
    '''
    Prints a pandas dataframe to an open excel 
    worksheet beginning at (startrow, startcol)

    Params
    ------
        df: dataframe to print

        ws: open worksheet

        header: True -> column headers will be 
        included starting one column to the right.

        index: True -> index will be included, 
        starting one row below the header.

        color: True -> rows will be colored blue;
        False -> no color
    '''
    rows = dataframe_to_rows(df, header=header, index=index)

    for r_idx, row in enumerate(rows, startrow):
        for c_idx, value in enumerate(row, startcol):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value
            if (color):
                cell.fill = blue_fill
            cell.border = border_style  # must be after color

def getDefaultPaths(config_file_path: str):
    try:
        with open(config_file_path, 'r') as file:
            directory_paths = json.load(file)

            time_forecast_directory = directory_paths.get("time_forecast_directory", "")
            report_directory = directory_paths.get("report_directory", "")
            contracts_list = directory_paths.get("contracts_list_filepath", "")
            team_members_list = directory_paths.get("team_members_list_filepath", "")

            return time_forecast_directory, report_directory, contracts_list, team_members_list

    except FileNotFoundError:
        print(f"Config file '{config_file_path}' not found.")
        return None

def getTeamList(PATH: str) -> pd.DataFrame:
    print("Reading TeamMembersList.xlsx...")
    try:
        team_list = pd.read_excel(PATH, "Sheet1", header=0)
    except Exception as e:
        print(e)
        return pd.DataFrame([])

    col_names = ['name', 'group', 'group_list', 'manager']
    team_list = team_list.set_axis(col_names, axis='columns')
    return team_list